import json
import os.path
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.serializers import serialize
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, JSONParser, FormParser
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from scenic.utils.pagination import Pagination
from scenic.models import *
from scenic.form.form import AttractionModelForm, UpForm, UpModelForm
from scenic.serializers import *
from openpyxl import load_workbook

""" drf学习使用 """
# class AttractionView(APIView):
#     def get(self, request):
#         attractions = Attraction.objects.all()
#         ser = AttractionSerializer(instance=attractions, many=True)
#         return Response(ser.data)
#
#     def post(self, request):
#         ser = AttractionSerializer(data=request.data)
#         if ser.is_valid():
#             ser.save()
#             return Response(ser.data)
#         else:
#             return Response(ser.errors)
#
#
# class AttractionDetailView(APIView):
#     def get(self, request, id):
#         attraction = Attraction.objects.get(pk=id)
#         ser = AttractionSerializer(instance=attraction, many=False)
#         return Response(ser.data)
#
#     def delete(self, request, id):
#         Attraction.objects.get(pk=id).delete()
#         return Response()
#
#     def put(self, request, id):
#         instance = Attraction.objects.get(pk=id)
#         ser = AttractionSerializer(data=request.data, instance=instance)
#         if ser.is_valid():
#             ser.save()
#             return Response(ser.data)
#         else:
#             return Response(ser.errors)


# Create your views here.
# class AttractionView(ListCreateAPIView):
#     queryset = Attraction.objects.all()
#     serializer_class = AttractionSerializer
#
#
# class AttractionDetailView(RetrieveUpdateDestroyAPIView):
#     queryset = Attraction.objects.all()
#     serializer_class = AttractionSerializer


class AttractionPageNumberPagination(PageNumberPagination):
    page_size = 10

import logging

logger = logging.getLogger(__name__)


class AttractionViewSet(ModelViewSet):
    queryset = Attraction.objects.all()
    serializer_class = AttractionSerializer
    pagination_class = AttractionPageNumberPagination
    parser_classes = [MultiPartParser, JSONParser, FormParser]

    def get_queryset(self):
        # 获取所有请求参数
        queryset = self.queryset
        scenic_id = self.request.query_params.get('scenic', None)
        search = self.request.query_params.get('search', None)
        # page = self.request.query_params.get('page')
        # 根据参数过滤queryset
        if scenic_id:
            queryset = queryset.filter(scenic=scenic_id)
        if search:
            # 添加基于 search 参数的过滤逻辑
            queryset = queryset.filter(attraction_name__icontains=search)  # 假设景点有一个name字段
        # 处理分页（如果启用了分页）
        # if page:
        #     paginator = self.pagination_class()
        #     result_page = paginator.paginate_queryset(queryset, page)
        #     return result_page
        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)  # 首先验证数据有效性
        # image = request.FILES.get('image')
        # if image:
        #     serializer.validated_data['image'] = image
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@csrf_exempt  # 没加这个会报错Forbidden (CSRF cookie not set.)
def upload(request):
    if request.method == 'POST' and request.FILES['image']:
        image = request.FILES['image']
        # 保存图片到指定目录
        image_path = os.path.join(settings.MEDIA_ROOT, 'attraction_images', image.name)
        with open(image_path, 'wb+') as destination:
            for chunk in image.chunks():
                destination.write(chunk)
        # 返回图片的 URL
        image_url = os.path.join(settings.MEDIA_URL, 'attraction_images', image.name)
        return JsonResponse({'success': True, 'url': image_url})
    return JsonResponse({'success': False, 'error': 'No image uploaded'})


def update_attraction_flow(request):
    if request.method == 'POST':
        # 从请求中获取用户位置信息和景点ID
        user_lat = request.POST.get('latitude')
        user_lng = request.POST.get('longitude')
        attraction_id = request.POST.get('attraction_id')

        # 查询指定ID的景点信息
        attraction = Attraction.objects.get(attraction_id=attraction_id)

        # 获取景点中心经纬度
        attraction_lat = attraction.attraction_lat
        attraction_lng = attraction.attraction_lng

        # 计算用户距离景点的距离（这里简化为直线距离）
        distance = calculate_distance(user_lat, user_lng, attraction_lat, attraction_lng)

        # 如果用户距离景点的距离在1km以内，更新景点的实时人流量信息
        if distance <= 1:
            # 假设每个用户的人流量都是1
            attraction.count += 1
            attraction.save()
            return JsonResponse({'message': 'User location updated successfully.'})
        else:
            # 用户离开景点范围，减少景点的实时人流量
            if attraction.count > 0:
                attraction.count -= 1
                attraction.save()

            return JsonResponse({'message': 'User left the attraction area.'})
    return JsonResponse({'error': 'Invalid request method or missing parameters.'}, status=400)


def calculate_distance(lat1, lng1, lat2, lng2):
    # 在这里编写计算两点距离的函数，可以参考你之前在前端代码中实现的 calculateDistance 函数
    pass


def get_ticket_attractions(request):
    attractions = Attraction.objects.filter(fee__isnull=False, fee__gt=0).values('attraction_id', 'attraction_name', 'fee', 'address', 'category', 'image')
    scenic_id = request.GET.get('scenic', None)
    search = request.GET.get('search', None)
    # 根据参数过滤queryset
    if scenic_id:
        attractions = attractions.filter(scenic=scenic_id)
    if search:
        # 添加基于 search 参数的过滤逻辑
        attractions = attractions.filter(attraction_name__icontains=search)
    attractions_list = list(attractions)
    return JsonResponse({'attractions': attractions_list})


def get_attraction(request):
    attractions = Attraction.objects.all()
    scenic_id = request.GET.get('scenic', None)
    search = request.GET.get('search', None)
    if scenic_id:
        attractions = attractions.filter(scenic=scenic_id)
    if search:
        # 添加基于 search 参数的过滤逻辑
        attractions = attractions.filter(attraction_name__icontains=search)
    data = [{'attraction_id': attraction.attraction_id,
             'scenic_id': attraction.scenic_id,
             'scenic_name': ScenicID(attraction.scenic_id).label,
             'attraction_name': attraction.attraction_name,
             'attraction_lng': attraction.attraction_lng,
             'attraction_lat': attraction.attraction_lat,
             'address': attraction.address,
             'description': attraction.description,
             'category': attraction.category,
             'fee': attraction.fee,
             'open_time': attraction.open_time,
             'close_time': attraction.close_time,
             'flow_limit': attraction.flow_limit,
             'status': attraction.status,
             'count': attraction.count,
             } for attraction in attractions]
    return JsonResponse(data, safe=False)

def get_attractions_list(request):
    # queryset = Attraction.objects.all()
    # paginator = AttractionPagination()
    # result_page = paginator.paginate_queryset(queryset, request)
    # serializer = AttractionSerializer(result_page, many=True)
    # return paginator.get_paginated_response(serializer.data)
    attractions_list = Attraction.objects.all()
    paginator = Paginator(attractions_list, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # attractions = list(page_obj.object_list.values())  # Convert queryset to list of dictionaries
    # 将每个景点的 scenic_id 转换为对应的景区名称，并添加到数据中
    attractions = []
    for attraction in page_obj.object_list:
        attraction_data = {
            'attraction_id': attraction.attraction_id,
            'scenic_id': attraction.scenic_id,
            'scenic_name': ScenicID(attraction.scenic_id).label,  # 获取景区名称
            'attraction_name': attraction.attraction_name,
            'attraction_lng': attraction.attraction_lng,
            'attraction_lat': attraction.attraction_lat,
            'address': attraction.address,
            'description': attraction.description,
            'category': attraction.category,
            'fee': attraction.fee,
            'open_time': attraction.open_time,
            'close_time': attraction.close_time,
            'flow_limit': attraction.flow_limit,
            'status': attraction.status,
            'count': attraction.count,
            'phone': attraction.phone,
        }
        attractions.append(attraction_data)
    return JsonResponse({'attractions': attractions, 'total_pages': paginator.num_pages})

def attraction_list(request):
    data_dict = {}
    search_data = request.GET.get('search', '')
    filter_scenic = request.GET.get('scenic', '')
    print(filter_scenic)
    if search_data:
        data_dict["attraction_name__contains"] = search_data
    if filter_scenic:
        queryset = Attraction.objects.filter(scenic=filter_scenic)
    else:
        queryset = Attraction.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)
    print(page_object.page_queryset)
    # queryset_serialized = serialize('python', page_object.page_queryset)  # 序列化 QuerySet 对象
    # queryset_data = [item['fields'] for item in queryset_serialized]  # 提取字典中的 'fields' 部分
    queryset_data = []
    for attraction in page_object.page_queryset:
        data = {
            'attraction_id': attraction.attraction_id,
            'scenic_name': attraction.scenic.scenic_name,  # 使用关联的景区名称替代景区id
            'attraction_name': attraction.attraction_name,
            'address': attraction.address,
            'description': attraction.description,
            'category': attraction.get_category_display(),
            'status': attraction.get_status_display(),
            'flow_limit': attraction.flow_limit,
            'phone': attraction.phone
            # 其他字段...
        }
        queryset_data.append(data)
    context = {
        # 'queryset': page_object.page_queryset,
        'queryset': queryset_data,
        'search_data': search_data,
        'select_scenic': filter_scenic,
        'page_string': page_object.html()
    }
    # return render(request, 'attraction_list.html', context)
    return JsonResponse(context)

# 手动写上传文件
def attraction_multi(request):
    """批量导入景点数据Excel"""
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 循环获取每行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = Attraction.objects.filter(attraction_name=text).exists()
        if not exists:
            Attraction.objects.create(attraction_name=text)
    return redirect('/api/attractionlist/')


# def upload_form(request):
#     if request.method == "GET":
#         form = UpForm()
#         return render(request, 'attraction_edit.html', {"form": form})
#     form = UpForm(data=request.POST, files=request.FILES)
#     if form.is_valid():
#         image_object = form.cleaned_data.get("img")
#         # db_file_path = os.path.join("static", "img", image_object.name)
#         media_path = os.path.join("media", image_object.name)
#         file_path = os.path.join("scenic", media_path)
#         f = open(file_path, mode='wb')
#         for chunk in image_object.chunks():
#             f.write(chunk)
#         f.close()
#
#         Attraction.objects.create(
#             attraction_name=form.cleaned_data['attraction_name'],
#             attraction_img=media_path
#         )

# ajax

# 上传文件建议用modelform，得配置media文件夹
def upload_model_form(request):
    if request.method == "GET":
        form = UpModelForm()
        return render(request, 'attraction_edit.html', {"form": form})
    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        form.save()


def attraction_detail(request):
    uid = request.GET.get("uid")
    row_object = Attraction.objects.filter(attraction_id=uid).first()
    if not row_object:
        return JsonResponse({"status": False, 'error': "数据不存在"})



def attraction_add(request):
    """原始方法"""
    if request.method == "GET":
        context = {
            "category_choices": Attraction.category_choices,
            "scenic_list": Scenic.objects.all()
        }
        return render(request, 'attraction_add.html', context)
    name = request.POST.get("name")
    desc = request.POST.get("desc")
    Attraction.objects.create(id=id, name=name, description=desc)
    return redirect("/api/attractionlist/")

def attraction_edit(request, nid):
    """编辑景点信息"""
    row_object = Attraction.objects.filter(attraction_id=nid).first()
    if request.method == "GET":
        form = AttractionModelForm(instance=row_object)
        return render(request, 'attraction_edit.html', {"form": form})
    form = AttractionModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/api/attractionlist/")
    return render(request, "attraction_edit.html", {"form": form})

def attraction_delete(request):
    nid = request.GET.get('nid')
    Attraction.objects.filter(attraction_id=nid).delete()
    return redirect("/api/attractionlist/")
